import sqlite3
import openpyxl
import re
import sys
import os
from bs4 import BeautifulSoup

sys.stdout.reconfigure(encoding='utf-8')

DB_PATH = "backend/data/scriptures.db"
EXCEL_PATH = "raw_data/Patanjali-yogasutra.IGS.xlsx"
HTML_PATH = "scratch/yogasuutra.html"

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def clean_text(s):
    if not s:
        return ''
    s = str(s).replace('_x0002_', '-')
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def clean_devanagari(s):
    if not s:
        return ''
    # Strip parenthetical annotations or prefixes/suffixes like (सर्वज्ञत्वबीजम्)
    s = re.sub(r'^\s*\(.*?\)\s*', '', s)
    s = re.sub(r'\s*\(.*?\)\s*$', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def parse_sanskrit_html():
    """Parses unaccented Devanagari Sanskrit verses from yogasuutra.html if available."""
    if not os.path.exists(HTML_PATH):
        return {}

    with open(HTML_PATH, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()
    
    # We use the unaccented section if present, or strip accents
    soup = BeautifulSoup(html, 'html.parser')
    text = soup.get_text()

    # If निःस्वर section exists, use that section
    if "निःस्वर" in text:
        text = text.split("निःस्वर")[-1]

    # Map Devanagari digits to ASCII
    deva_to_ascii = str.maketrans('०१२३४५६७८९', '0123456789')
    text = text.translate(deva_to_ascii)

    chapters_sanskrit = {1: {}, 2: {}, 3: {}, 4: {}}
    matches = re.finditer(r'(?:॥.*?॥)?(.*?)\s*॥\s*(\d+)\.(\d+)\s*॥', text.replace('\n', ' '))
    for m in matches:
        sanskrit = m.group(1).strip()
        sanskrit = re.sub(r'.*॥\s*', '', sanskrit).strip()
        sanskrit = re.sub(r'[॒॑]', '', sanskrit)
        sanskrit = clean_devanagari(sanskrit)

        ch = int(m.group(2))
        v = int(m.group(3))

        if 1 <= ch <= 4:
            chapters_sanskrit[ch][v] = sanskrit

    return chapters_sanskrit

def load_excel_data():
    print(f"Loading Excel data from {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb.active

    parsed = {1: [], 2: [], 3: [], 4: []}
    current_ch = 0

    for r in range(1, sheet.max_row + 1):
        c0 = sheet.cell(r, 1).value
        c1 = sheet.cell(r, 2).value
        c2 = sheet.cell(r, 3).value
        c3 = sheet.cell(r, 4).value

        if c0 and 'CHAPTER' in str(c0):
            m = re.search(r'CHAPTER\s*(\d+)', str(c0), re.IGNORECASE)
            if m:
                current_ch = int(m.group(1))

        if current_ch > 0 and (c1 is not None or c2 is not None or c3 is not None):
            if c0 and 'CHAPTER' in str(c0) and c1 is None:
                continue
            v_num = int(c1) if c1 is not None else len(parsed[current_ch]) + 1
            translit = clean_text(c2)
            meaning = clean_text(c3)

            # Fix known cell splicing anomalies from raw Excel
            if 'From direct intuitive perception' in translit:
                translit = translit.replace('From direct intuitive perception', '').strip()
                meaning = 'From direct intuitive perception ' + meaning

            if 'When there is equal purity' in translit:
                translit = translit.replace('When there is equal purity', '').strip()
                meaning = 'When there is equal purity ' + meaning

            parsed[current_ch].append({
                'chapter': current_ch,
                'verse_number': v_num,
                'transliteration': translit,
                'meaning': meaning
            })

    return parsed

def get_complete_dataset():
    excel_data = load_excel_data()
    sanskrit_data = parse_sanskrit_html()

    # Fallback to existing DB if html missing
    conn = get_db()
    c = conn.cursor()
    db_sanskrit = {1: {}, 2: {}, 3: {}, 4: {}}
    for ch in range(1, 5):
        c.execute("""
            SELECT v.verse_number, v.sanskrit_text
            FROM verses v
            JOIN sections s ON v.section_id = s.id
            JOIN sources src ON s.source_id = src.id
            WHERE src.name = 'Patanjali Yoga Sutras' AND s.chapter_number = ?
            ORDER BY v.verse_number
        """, (ch,))
        for r in c.fetchall():
            db_sanskrit[ch][r[0]] = clean_devanagari(r[1])
    conn.close()

    chapter_names = {
        1: "Samadhi Pada",
        2: "Sadhana Pada",
        3: "Vibhuti Pada",
        4: "Kaivalya Pada"
    }

    dataset = []
    for ch in range(1, 5):
        verses_list = []
        for item in excel_data[ch]:
            v_num = item['verse_number']
            translit = item['transliteration']
            meaning = item['meaning']

            # Devanagari Sanskrit mapping
            sanskrit = ""
            if ch == 1:
                sanskrit = sanskrit_data.get(1, {}).get(v_num) or db_sanskrit[1].get(v_num, "")
            elif ch == 2:
                sanskrit = sanskrit_data.get(2, {}).get(v_num) or db_sanskrit[2].get(v_num, "")
            elif ch == 3:
                if v_num <= 21:
                    sanskrit = sanskrit_data.get(3, {}).get(v_num) or db_sanskrit[3].get(v_num, "")
                elif v_num == 22:
                    sanskrit = "एतेन शब्दाद्यन्तर्धानमुक्तम्"
                else:
                    # shifted by 1 due to 3.22
                    sanskrit = sanskrit_data.get(3, {}).get(v_num - 1) or db_sanskrit[3].get(v_num - 1, "")
            elif ch == 4:
                sanskrit = sanskrit_data.get(4, {}).get(v_num) or db_sanskrit[4].get(v_num, "")

            sanskrit = clean_devanagari(sanskrit)

            verses_list.append({
                'verse_number': v_num,
                'sanskrit_text': sanskrit,
                'transliteration': translit,
                'meaning': meaning
            })

        dataset.append({
            'chapter_number': ch,
            'chapter_name': chapter_names[ch],
            'verses': verses_list
        })

    return dataset

def ingest():
    dataset = get_complete_dataset()
    conn = get_db()
    cursor = conn.cursor()

    print("Checking / Creating source entry in database...")
    cursor.execute("SELECT id FROM sources WHERE name = ?", ("Patanjali Yoga Sutras",))
    source_row = cursor.fetchone()
    if source_row:
        source_id = source_row[0]
        print(f"Found existing source_id: {source_id}")
    else:
        cursor.execute("INSERT INTO sources (name, type) VALUES (?, ?)", ("Patanjali Yoga Sutras", "Sutra"))
        source_id = cursor.lastrowid
        print(f"Created new source_id: {source_id}")

    # Remove previous sections, verses, and translations for Patanjali Yoga Sutras
    cursor.execute("""
        SELECT id FROM sections WHERE source_id = ?
    """, (source_id,))
    old_section_ids = [r[0] for r in cursor.fetchall()]
    
    if old_section_ids:
        placeholders = ','.join('?' * len(old_section_ids))
        cursor.execute(f"""
            SELECT id FROM verses WHERE section_id IN ({placeholders})
        """, old_section_ids)
        old_verse_ids = [r[0] for r in cursor.fetchall()]

        if old_verse_ids:
            v_placeholders = ','.join('?' * len(old_verse_ids))
            cursor.execute(f"DELETE FROM translations WHERE verse_id IN ({v_placeholders})", old_verse_ids)
            cursor.execute(f"DELETE FROM commentaries WHERE verse_id IN ({v_placeholders})", old_verse_ids)
            cursor.execute(f"DELETE FROM verses WHERE id IN ({v_placeholders})", old_verse_ids)

        cursor.execute(f"DELETE FROM sections WHERE id IN ({placeholders})", old_section_ids)

    print("Inserting fresh sections, verses, transliterations, and IGS translations...")
    total_verses = 0
    for ch_data in dataset:
        ch_num = ch_data["chapter_number"]
        ch_name = ch_data["chapter_name"]

        cursor.execute("""
            INSERT INTO sections (source_id, chapter_number, chapter_name)
            VALUES (?, ?, ?)
        """, (source_id, ch_num, ch_name))
        section_id = cursor.lastrowid

        for v in ch_data["verses"]:
            v_num = v["verse_number"]
            sanskrit = v["sanskrit_text"]
            translit = v["transliteration"]
            meaning = v["meaning"]

            cursor.execute("""
                INSERT INTO verses (section_id, verse_number, sanskrit_text, transliteration)
                VALUES (?, ?, ?, ?)
            """, (section_id, v_num, sanskrit, translit))
            verse_id = cursor.lastrowid

            if meaning:
                cursor.execute("""
                    INSERT INTO translations (verse_id, language, text, author)
                    VALUES (?, ?, ?, ?)
                """, (verse_id, "english", meaning, "International Gita Society"))

            total_verses += 1

    conn.commit()
    conn.close()
    print(f"Successfully ingested {total_verses} Yoga Sutras verses with IGS translations!")

if __name__ == "__main__":
    ingest()
